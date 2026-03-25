"""
src/ledger/infrastructure/parsers.py
====================================
Unified parsing logic for PDF, CSV, and Excel documents.
Supports text extraction and structure-aware parsing.
"""

import csv
from pathlib import Path
from typing import Any

import openpyxl
from pypdf import PdfReader


class DocumentParserError(Exception):
    """Base error for document parsing issues."""

    pass


def parse_pdf(file_path: str) -> str:
    """Extraxt text from all pages of a PDF file."""
    try:
        reader = PdfReader(file_path)
        text = []
        for page in reader.pages:
            content = page.extract_text()
            if content:
                text.append(content)
        return "\n\n--- Page Break ---\n\n".join(text)
    except Exception as e:
        raise DocumentParserError(f"Failed to parse PDF: {e}") from e


def parse_csv(file_path: str) -> list[dict[str, Any]]:
    """Parse a CSV file into a list of dictionaries."""
    try:
        with Path(file_path).open(encoding="utf-8") as f:
            reader = csv.DictReader(f)
            return list(reader)
    except Exception as e:
        raise DocumentParserError(f"Failed to parse CSV: {e}") from e


def parse_excel(file_path: str) -> dict[str, list[list[Any]]]:
    """Parse an Excel file into a dictionary of sheets, where each sheet is a list of rows."""
    try:
        wb = (
            openpyxl.load_model(file_path)
            if hasattr(openpyxl, "load_model")
            else openpyxl.load_workbook(file_path, data_only=True)
        )
        # Fallback to load_workbook if load_model doesn't exist
        # (it usually doesn't in standard openpyxl)
        if not hasattr(openpyxl, "load_model"):
            wb = openpyxl.load_workbook(file_path, data_only=True)

        sheets = {}
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            data = []
            for row in sheet.iter_rows(values_only=True):
                data.append(list(row))
            sheets[sheet_name] = data
        return sheets
    except Exception as e:
        raise DocumentParserError(f"Failed to parse Excel: {e}") from e


def parse_any(file_path: str) -> Any:
    """Detect file type and parse accordingly."""
    suffix = Path(file_path).suffix.lower()
    if suffix == ".pdf":
        return parse_pdf(file_path)
    if suffix == ".csv":
        return parse_csv(file_path)
    if suffix in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        return parse_excel(file_path)
    if suffix in (".txt", ".md", ".json", ".jsonl"):
        with Path(file_path).open(encoding="utf-8") as f:
            return f.read()
    else:
        raise DocumentParserError(f"Unsupported file type: {suffix}")
